import os
import pandas as pd
from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import easygui
import xlwings as xw

#### Indicate the version of the SEALigHTS analysis pipeline:
version_pipeline = '_UMI_V4'

READ_ME = """ 
Critieria used for the search of splicing anomalies in V4:
- highly expressed junction definition ("Jonction fortes") : linear canonical splicing junction with a median count > 50UMI
- Exonskipping :
                - Normalised count of the junction >q75+1.5IQR  
                - Normalised count of the junction > 1/25 of the q50 of the canonical junctions inside the skipping (1/50 for PMS2)
                - fold change > 2
                - more than 10% of canonical junctions inside the jump are < q25
                - junction excluded if count < 5 UMI 
- Dup_orcircRNA:
                - Normalised count of the junction >q75+1.5IQR 
                - Normalised count of the junction > 1/3 of the q50 of the canonical junctions inside the circRNA or duplication
                - if q50 =0 : the anomaly "Type" is named Dup+++_or_circRNA
- low junction:
                - Normalised count of the junction < q25 - 1.5IQR
                - Normalised count of the junction < 75% of the q50 of this junction
                - the juncton must be categorised as a "Jonction_forte"
                - Score 11 = ratio of the junction compared to the upstream and downstream junctions is < q25-1.5IQR
                - Score 10 = only the ratio with the upstream junction is < q25
                - Score 01 = only the ratio with the downstream junction is < q25
- Low_gene_expression: 
                Calculation of the ratio of each junction of a gene between the patient and the median for the run 
                - the median of each ratio must be < 2/3 (66%)
                - more than 80% of the junctions must be below < q25
                            
"""


gene_dicts = {
    "Digestive panel": ['APC', 'BMPR1A', 'EPCAM', 'MLH1', 'MSH2', 'MSH6', 'NTHL1', 'MUTYH', 'PMS2', 'POLD1', 'POLE','SMAD4', 'STK11', 'CDH1', 'MSH3', 'AXIN2', 'GREM1', 'RNF43', 'GALNT12', 'BUB1', 'RPS20', 'FAN1',],
    "HBOC panel": ['BRCA1', 'BRCA2', 'PALB2', 'RAD51C', 'RAD51D', 'CDH1', 'TP53', 'PTEN','MLH1', 'MSH2', 'MSH6', 'PMS2','ATM', 'BAP1', 'BARD1', 'BRIP1', 'CHEK2', 'FAM175A', 'FANCM', 'MRE11', 'NBN', 'RAD50', 'RAD51B',
                       'RINT1', 'XRCC2']
    }

#---------------------Définition des fonctions-------------------------------
def select_file():
    file = easygui.fileopenbox(msg="Select the Result file you want to format", title="Result file to format",
                               default='*.csv', filetypes=['*.csv'])
    return file


def get_run_name(file):
    run_name = os.path.basename(os.path.dirname(os.path.dirname(file)))
    print("run : " + run_name)
    return run_name

def get_repertoire_name(file):
    rep = os.path.dirname(os.path.dirname(file))
    return rep

def open_file(path):
    if path.endswith('.csv'):
        df = pd.read_csv(path, sep=';')
    else:
        print('Error : file format not recognized, csv expected')
    return df


def add_links(df):
    df.insert(3, 'Mean_Figure', '')

    def generate_link_Mean(row):
        gene = row['Gene']
        link = os.path.join('Mean', str(gene), 'Figures', 'Mean_Splice_' + str(gene) + '.png')
        return '=HYPERLINK("' + link + '", "Mean_Figure")'

    df['Mean_Figure'] = df.apply(generate_link_Mean, axis=1)

    df.insert(4, 'Patient_figure', '')

    def generate_link_patient(row):
        gene = row['Gene']
        patient = row['Patient']
        link = os.path.join(str(patient), str(gene), 'Figures', str(patient) + '_Splice_' + str(gene) + '.png')
        return '=HYPERLINK("' + link + '", "Patient_figure")'

    df['Patient_figure'] = df.apply(generate_link_patient, axis=1)

    df.insert(16, 'Matrice UMI', '')

    def generate_link_matrice_UMI (row):
        gene = row['Gene']
        link = os.path.join('Matrices', 'Matrice_UMI_' + str(gene) + '_' + run + '.csv')
        return '=HYPERLINK("' + link + '", "Matrice UMI")'

    df['Matrice UMI'] = df.apply(generate_link_matrice_UMI, axis=1)

    df.insert(14, 'Matrice Full', '')

    def generate_link_matrice_Full(row):
        gene = row['Gene']

        link = os.path.join('Matrices', 'Matrice_Full_' + str(gene) + '_' +run + '.csv')
        return '=HYPERLINK("' + link + '", "Matrice Full")'
    df['Matrice Full'] = df.apply(generate_link_matrice_Full, axis=1)

    df.insert(18, 'Ano_Splice', '')


    def add_link_Ano_patient(row):
        gene = row['Gene']
        patient = row['Patient']
        link = os.path.join('Resultats', 'Ano_Splice_' + str(patient) + '.csv')
        return '=HYPERLINK("' + link + '", "Ano_Splice")'

    df ['Ano_Splice'] = df.apply(add_link_Ano_patient, axis=1)
    return df


def add_occurences_columns(df):
    # count the number of anomalies of each sample
    occurrences_patient = df['Patient'].value_counts()
    df['Patient occurences'] = df['Patient'].map(occurrences_patient)

    Alarme = df['Gene'] + df['Jonction'] + df['Type']

    # Count the number of times an anomaly occurs in the run 
    occurrences_Alarme = Alarme.value_counts()
    df['Anomaly occurences'] = Alarme.map(occurrences_Alarme)
    return df

#add a column indicating the panel of the gene:
def add_gene_list_names(df):
    def get_gene_list_names(gene):
        result = []
        for gene_list_name, gene_list in gene_dicts.items():
            if gene in gene_list:
                result.append(gene_list_name)
        return ', '.join(result)

    df.insert(2, 'Panel', '') 
    df['Panel'] = df['Gene'].apply(
        get_gene_list_names)  
    return df


def pd_to_xl(df, input_csv):
    output_xl = input_csv.replace('.csv', '_mis_en_forme.xlsx')
    df.to_excel(output_xl, sheet_name='Feuille1', index=False)
    return output_xl


def format_xl_file(xl):
    wb = load_workbook(xl)
    sheet = wb['Feuille1']

    header_range = 'A1:T1'

    header_fill = PatternFill(start_color='D5D5D5', end_color='D5D5D5', fill_type='solid')
    for row in sheet[header_range]:
        for cell in row:
            cell.fill = header_fill

    # automaticaly adjusting column lenght
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # adjust specific column lenght
    sheet.column_dimensions['C'].width = 16
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 14
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['J'].width = 10
    sheet.column_dimensions['L'].width = 10
    sheet.column_dimensions['M'].width = 10
    sheet.column_dimensions['N'].width = 10
    sheet.column_dimensions['O'].width = 15
    sheet.column_dimensions['P'].width = 15
    sheet.column_dimensions['Q'].width = 17
    sheet.column_dimensions['S'].width = 15
    sheet.column_dimensions['T'].width = 15

    wb.save(xl)
    return xl

def round_to_1_decimals(df):
    df['Mean_UMI'] = pd.to_numeric(df['Mean_UMI'],errors='coerce')
    df['Mean_Tot_UMI'] = pd.to_numeric(df['Mean_Tot_UMI'], errors='coerce')
    df['Mean_UMI'] = df['Mean_UMI'].apply(lambda x: round(x, 1))
    df['Mean_Tot_UMI'] = df['Mean_Tot_UMI'].apply(lambda x: round(x, 1))
    return df

def open_xl(output_xl):
    wb = xw.Book(output_xl)
    wb.app.visible = True
    wb.app.activate()
    wb.sheets[0].range('A1').select()


# ---------------------------Begining of the script--------------------------

input_file = select_file()
df= open_file(input_file)
run = get_run_name(input_file)
rep = get_repertoire_name(input_file)
add_occurences_columns(df)
add_links(df)
df = add_gene_list_names(df)
df = round_to_1_decimals(df)

output_xl = rep + os.path.sep + 'Results_' + run + '_formatted.xlsx'
df.to_excel(output_xl, sheet_name='Feuille1', index=False)
format_xl_file(output_xl)


print('The result file has been formatted and is open in excel, it is saved in the folder out, run: ' + run )
#open automaticaly the output file
open_xl(output_xl)
